#!/usr/bin/env python3
# _*_ coding: utf-8 _*_
# @Time : 2022/11/28 19:18
# @Author : wdm
# @desc : dm_user表接口

from typing import Any, List
from fastapi import APIRouter, Depends, Request, File, UploadFile, Form
from openpyxl.styles import Alignment, Border, Side, Font
#from sqlalchemy import and_, or_
from sqlalchemy.orm import Session

from schemas import Dm_userUpdate, Dm_userCreate, Dm_userOut, Dm_userDelete, ResultModel, Dm_userType, Dm_userCodeCheck, Dm_userCodeCheckNoid, Dm_userPwd
from crud import dm_user, adminauthclass
from utils import resp_200, resp_400, IdNotExist, logger
from models import Dm_user, Admin, Adminauthclass

#import shutil
from core import settings
#from utils.create_dir import create_dir
#from pathlib import Path
#from tempfile import NamedTemporaryFile
from api.deps import get_db, get_current_user, checkOxygenPulseConclusion, checkHeartrateConclusion, checkTemperatureConclusion, \
    checkBloodPressureConclusion, checkGluPMBGConclusion, checkOxygenSpoConclusion, checkGluPBGConclusion, checkPhysiqueConclusion,\
    getUserDeviceWhere, getUserWhere

#from core import get_password_hash

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Series, Reference
#from openpyxl.chart.layout import Layout, ManualLayout
#from copy import deepcopy

import time

#import datetime

router = APIRouter()

@router.get("/", response_model=ResultModel[Dm_userOut], summary='查询所有用户号(根据页码和每页个数)')
def read_dm_users(db: Session = Depends(get_db), pageIndex: int = 1, pageSize: int = 10) -> Any:
    get_dm_users = dm_user.get_multi(db, pageIndex=pageIndex, pageSize=pageSize)
    return resp_200(data=get_dm_users, msg=f"查询了第 {pageIndex} 页中的 {pageSize} 个用户号信息.")

@router.get("/qtreebloodtype", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询血型信息')
def query_bloodtype_tree(bloodtype_name: str = "") -> Any:
    where = " where 1=1 "
    if str.strip(bloodtype_name) != "" :
        where = where + " and (name like '%%" + str(bloodtype_name) + "%%')"
    sql = "select id as bloodtypeid,name as bloodtypename,seq from b_bloodtype " + where
    get_bloodtypes = dm_user.get_by_sql_no_count(sql)
    if not get_bloodtypes:
        raise IdNotExist(f"系统中不存在 血型查询关键字 为 {bloodtype_name} 的血型结果.")
    return resp_200(data=get_bloodtypes, msg=f"查询到了血型结果.")

@router.get("/qtreeallergy", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询药物过敏史信息')
def query_allergy_tree(allergy_name: str = "") -> Any:
    where = " where 1=1 "
    if str.strip(allergy_name) != "" :
        where = where + " and (name like '%%" + str(allergy_name) + "%%')"
    sql = "select id as allergyid,name as allergyname,seq from b_allergy " + where
    get_allergys = dm_user.get_by_sql_no_count(sql)
    if not get_allergys:
        raise IdNotExist(f"系统中不存在 药物过敏史查询关键字 为 {allergy_name} 的药物过敏史结果.")
    return resp_200(data=get_allergys, msg=f"查询到了药物过敏史结果.")

@router.get("/qtreemedical", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询既往病史信息')
def query_medical_tree(medical_name: str = "") -> Any:
    where = " where 1=1 "
    if str.strip(medical_name) != "" :
        where = where + " and (name like '%%" + str(medical_name) + "%%')"
    sql = "select id as medicalid,name as medicalname,seq from b_medical " + where
    get_medicals = dm_user.get_by_sql_no_count(sql)
    if not get_medicals:
        raise IdNotExist(f"系统中不存在 既往病史查询关键字 为 {medical_name} 的既往病史结果.")
    return resp_200(data=get_medicals, msg=f"查询到了既往病史结果.")

@router.get("/qtreegenetic", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询遗传史信息')
def query_genetic_tree(genetic_name: str = "") -> Any:
    where = " where 1=1 "
    if str.strip(genetic_name) != "" :
        where = where + " and (name like '%%" + str(genetic_name) + "%%')"
    sql = "select id as geneticid,name as geneticname,seq from b_genetic " + where
    get_genetics = dm_user.get_by_sql_no_count(sql)
    if not get_genetics:
        raise IdNotExist(f"系统中不存在 遗传史查询关键字 为 {genetic_name} 的遗传史结果.")
    return resp_200(data=get_genetics, msg=f"查询到了遗传史结果.")

@router.post("/updatejson", response_model=ResultModel[Dm_userOut], summary='添加用户信息')
def update_dm_user_json(*, db: Session = Depends(get_db), dm_user_in: Dm_userCreate) -> Any:
    #print('dm_user_in', dm_user_in)
    if dm_user_in.id != 0:#编辑
        dm_user_update = dict()
        havecity,havebloodtype,haveallergy,havemedical,havegenetic = False,False,False,False,False
        for key,value in dm_user_in:
            if key != "id":
                # 赋值字典
                if key == "bloodtypename" or key == "allergyname" or key == "medicalname" or key == "geneticname":
                    '''if value != "":
                        pass
                    else:
                        dm_user_update[key] = value'''
                    pass
                else:
                    dm_user_update[key] = value
                where = " where 1=1 "
                bname = ""
                if key == "cityid":
                    havecity = True
                if key == "bloodtypeid":
                    havebloodtype = True
                    if value != 0:
                        where = where + " and id = " + str(value)
                        sql = "select id as bloodtypeid,name as bloodtypename from b_bloodtype " + where
                        get_bloodtypes = dm_user.get_by_sql_no_count(sql)
                        if get_bloodtypes:
                            for k, v in get_bloodtypes:
                                bname = v
                            dm_user_update["bloodtypename"] = bname
                if key == "allergyid":
                    haveallergy = True
                    if value != 0:
                        where = where + " and id = " + str(value)
                        sql = "select id as allergyid,name as allergyname from b_allergy " + where
                        get_allergys = dm_user.get_by_sql_no_count(sql)
                        if get_allergys:
                            for k, v in get_allergys:
                                bname = v
                            dm_user_update["allergyname"] = bname
                if key == "medicalid":
                    havemedical = True
                    if value != 0:
                        where = where + " and id = " + str(value)
                        sql = "select id as medicalid,name as medicalname from b_medical " + where
                        get_medicals = dm_user.get_by_sql_no_count(sql)
                        if get_medicals:
                            for k, v in get_medicals:
                                bname = v
                            dm_user_update["medicalname"] = bname
                if key == "geneticid":
                    havegenetic = True
                    if value != 0:
                        where = where + " and id = " + str(value)
                        sql = "select id as geneticid,name as geneticname from b_genetic " + where
                        get_genetics = dm_user.get_by_sql_no_count(sql)
                        if get_genetics:
                            for k, v in get_genetics:
                                bname = v
                            dm_user_update["geneticname"] = bname
        if havecity == False:
            dm_user_update['cityid'] = 0
        if havebloodtype == False:
            dm_user_update['bloodtypeid'] = 0
        if haveallergy == False:
            dm_user_update['allergyid'] = 0
        if havemedical == False:
            dm_user_update['medicalid'] = 0
        if havegenetic == False:
            dm_user_update['geneticid'] = 0
        get_dm_user = dm_user.get(db, id=dm_user_in.id)
        print(dm_user_update)
        alter_dm_user = dm_user.update(db, db_obj=get_dm_user, obj_in=dm_user_update)
        return resp_200(data=alter_dm_user, msg=f"更新了 id 为 {id} 的用户信息.")
    else:#新增
        dm_user_add = dict()
        havecity,havebloodtype,haveallergy,havemedical,havegenetic = False,False,False,False,False
        for key,value in dm_user_in:
            if key != "id":
                # 赋值字典
                if key == "bloodtypename" or key == "allergyname" or key == "medicalname" or key == "geneticname":
                    if value != "":
                        pass
                    else:
                        dm_user_add[key] = value
                else:
                    dm_user_add[key] = value
                where = " where 1=1 "
                bname = ""
                if key == "cityid":
                    havecity = True
                if key == "bloodtypeid":
                    havebloodtype = True
                    if value != 0:
                        where = where + " and id = " + str(value)
                        sql = "select id as bloodtypeid,name as bloodtypename from b_bloodtype " + where
                        get_bloodtypes = dm_user.get_by_sql_no_count(sql)
                        if get_bloodtypes:
                            for k, v in get_bloodtypes:
                                bname = v
                            dm_user_add["bloodtypename"] = bname
                if key == "allergyid":
                    haveallergy = True
                    if value != 0:
                        where = where + " and id = " + str(value)
                        sql = "select id as allergyid,name as allergyname from b_allergy " + where
                        get_allergys = dm_user.get_by_sql_no_count(sql)
                        if get_allergys:
                            for k, v in get_allergys:
                                bname = v
                            dm_user_add["allergyname"] = bname
                if key == "medicalid":
                    havemedical = True
                    if value != 0:
                        where = where + " and id = " + str(value)
                        sql = "select id as medicalid,name as medicalname from b_medical " + where
                        get_medicals = dm_user.get_by_sql_no_count(sql)
                        if get_medicals:
                            for k, v in get_medicals:
                                bname = v
                            dm_user_add["medicalname"] = bname
                if key == "geneticid":
                    havegenetic = True
                    if value != 0:
                        where = where + " and id = " + str(value)
                        sql = "select id as geneticid,name as geneticname from b_genetic " + where
                        get_genetics = dm_user.get_by_sql_no_count(sql)
                        if get_genetics:
                            for k, v in get_genetics:
                                bname = v
                            dm_user_add["geneticname"] = bname
        if havecity == False:
            dm_user_add['cityid'] = 0
        if havebloodtype == False:
            dm_user_add['bloodtypeid'] = 0
        if haveallergy == False:
            dm_user_add['allergyid'] = 0
        if havemedical == False:
            dm_user_add['medicalid'] = 0
        if havegenetic == False:
            dm_user_add['geneticid'] = 0
        add_dm_user = dm_user.create(db, obj_in=dm_user_add)
        return resp_200(data=add_dm_user, msg=f"添加了 id 为 {add_dm_user.id} 的用户信息.")

@router.post("/settype", response_model=ResultModel[Dm_userOut], summary='添加dm_user信息')
def update_dm_user_type(*, db: Session = Depends(get_db), dm_user_in: Dm_userType) -> Any:
    if dm_user_in.id == 0:
        raise IdNotExist(f"缺少id.")
    if dm_user_in.is_active == 0 or dm_user_in.is_active == 1:
        get_dm_user = dm_user.get(db, id=dm_user_in.id)
        alter_dm_user = dm_user.update(db, db_obj=get_dm_user, obj_in=dm_user_in)
        return resp_200(data=alter_dm_user, msg=f"更新了 id 为 {dm_user_in.id} 的dm_user激活信息.")
    else:
        raise IdNotExist(f"缺少参数.")

@router.delete("/delete", response_model=ResultModel[Dm_userOut], summary='通过 id 删除用户信息')
def delete_dm_user(*, db: Session = Depends(get_db), dm_user_id: Dm_userDelete) -> Any:
    del_dm_user = dm_user.remove(db, id=dm_user_id.id)
    #print('\n'.join(['{0}: {1}'.format(item[0], item[1]) for item in del_dm_user.__dict__.items()]))
    return resp_200(data=del_dm_user, msg=f"删除了 id 为 {dm_user_id.id} 的用户信息.")

@router.get("/querybyid", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询dm_user信息')
def query_dm_user_id(
        user_id: int = 0, db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
     ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserDeviceWhere(db,current_user)
    where += strWhereUser
    if user_id != 0 :
        where = where + " and A.id=" + str(user_id)
        sql = "select A.id,A.name,A.headicon,A.age,A.sex,A.deviceid,DV.name as devicename,A.idcard,A.height,A.weight,A.cityid,A.cityname,A.phone"
        sql = sql  + " ,A.bloodtypeid,if(IsNULL(A1.name), A.bloodtypename, A1.name) as bloodtypename "
        sql = sql  + " ,A.allergyid,if(IsNULL(A2.name), A.allergyname, A2.name) as allergyname "
        sql = sql  + " ,A.medicalid,if(IsNULL(A3.name), A.medicalname, A3.name) as medicalname "
        sql = sql  + " ,A.geneticid,if(IsNULL(A4.name), A.geneticname, A4.name) as geneticname "
        sql = sql  + " ,A.usergeneticname,A.descr,A.createtime "
        sql = sql  + " from dm_user A left join device DV on A.deviceid=DV.device_code "
        sql = sql  + "  left join b_bloodtype A1 on A.bloodtypeid=A1.id "
        sql = sql  + "  left join b_allergy A2 on A.allergyid=A2.id "
        sql = sql  + "  left join b_medical A3 on A.medicalid=A3.id "
        sql = sql  + "  left join b_genetic A4 on A.geneticid=A4.id " + where
        get_dm_users = dm_user.get_by_sql_no_count(sql)
        if not get_dm_users:
            raise IdNotExist(f"系统中不存在 dm_user查询关键字 为 {user_id} 的dm_user结果.")
        return resp_200(data=get_dm_users, msg=f"查询到了 dm_user查询关键字 为 {user_id} 的dm_user结果.")
    else:
        return resp_400()

@router.get("/queryweb", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询dm_user信息')
def query_dm_user(
        user_name: str = "", deviceId: str = "", productId: str = "", deviceidlength: int = 0, pageIndex: int = 1, pageSize: int = 10,
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
        ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserDeviceWhere(db,current_user)
    where += strWhereUser
    if str.strip(user_name) != "" :
        where = where + " and (A.name like '%%" + str(user_name) + "%%'"
        where = where + " or A.descr like '%%" + str(user_name) + "%%')"
    if deviceId != "":
        where = where + " and A.deviceid like '%%" + str(deviceId) + "%%'"
    if productId != "":
        where = where + " and A.deviceid in (select device_code as deviceid from device where product_id = " + str(productId) + ")"
    if pageIndex<1:
        raise IdNotExist(f"页数小于1.")
    if pageSize<1:
        raise IdNotExist(f"每页记录数小于1.")
    if deviceidlength != 0:
        where = where + " and length(A.deviceid) =" + str(deviceidlength)
    order = " order by A.createtime desc,A.deviceid"
    limit = " limit " + str((pageIndex - 1) * pageSize) + "," + str(pageSize)
    sql = "select A.id,A.name,A.headicon,A.age,A.sex,A.deviceid,DV.name as devicename,A.idcard,A.height,A.weight,A.cityid,A.cityname,A.phone"
    sql = sql  + " ,A.bloodtypeid,if(IsNULL(A1.name), A.bloodtypename, A1.name) as bloodtypename "
    sql = sql  + " ,A.allergyid,if(IsNULL(A2.name), A.allergyname, A2.name) as allergyname "
    sql = sql  + " ,A.medicalid,if(IsNULL(A3.name), A.medicalname, A3.name) as medicalname "
    sql = sql  + " ,A.geneticid,if(IsNULL(A4.name), A.geneticname, A4.name) as geneticname "
    sql = sql  + " ,A.usergeneticname,A.descr,A.createtime "
    sql = sql  + " ,B1.bloodsugarcount,B2.bloodpressurecount,B3.heartratecount,B4.bloodoxygencount,B5.temperaturecount "
    sql = sql  + " from dm_user A left join device DV on A.deviceid=DV.device_code "
    sql = sql  + "  left join b_bloodtype A1 on A.bloodtypeid=A1.id "
    sql = sql  + "  left join b_allergy A2 on A.allergyid=A2.id "
    sql = sql  + "  left join b_medical A3 on A.medicalid=A3.id "
    sql = sql  + "  left join b_genetic A4 on A.geneticid=A4.id "
    sql = sql  + "  left join (select dmuserid,count(dmuserid) as bloodsugarcount from dm_pbloodsugar group by dmuserid) B1 on A.id=B1.dmuserid "
    sql = sql  + "  left join (select dmuserid,count(dmuserid) as bloodpressurecount from dm_pbloodpressure group by dmuserid) B2 on A.id=B2.dmuserid "
    sql = sql  + "  left join (select dmuserid,count(dmuserid) as heartratecount from dm_pheartrate group by dmuserid) B3 on A.id=B3.dmuserid "
    sql = sql  + "  left join (select dmuserid,count(dmuserid) as bloodoxygencount from dm_pbloodoxygen group by dmuserid) B4 on A.id=B4.dmuserid "
    sql = sql  + "  left join (select dmuserid,count(dmuserid) as temperaturecount from dm_ptemperature group by dmuserid) B5 on A.id=B5.dmuserid "
    sql = sql + where + order + limit
    #print(sql)
    sqlcount = "select count(A.id) as countid from dm_user A left join device DV on A.deviceid=DV.device_code " + where + order
    get_dm_users = dm_user.get_by_sql(sql, sqlcount)
    if not get_dm_users:
        raise IdNotExist(f"系统中不存在 dm_user查询关键字 为 {user_name} 的用户结果.")
    '''for result in get_dm_users:
        for k, v in result.items():
            if k == "dmuserid" :
                sql1 = "select count(A.id) as countid from dm_pbloodsugar A where 1=1 and dmuserid = " + str(v)
                kvcount = dm_user.get_by_sql_count_int(sql1)
                result.items()'''
    return resp_200(data=get_dm_users, msg=f"查询到了 dm_user查询关键字 为 {user_name} 的用户结果.")

@router.post("/query", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户号信息')
def query_dm_user(productId: int = 0, pageIndex: int = 1, pageSize: int = 10) -> Any:
    where = " where 1=1 "
    if productId != 0 :
        where = where + " and product_id = " + str(productId)
    else:
        raise IdNotExist(f"用户类ID必须.")
    if pageIndex<1:
        raise IdNotExist(f"页数小于1.")
    if pageSize<1:
        raise IdNotExist(f"每页记录数小于1.")
    limit = " limit " + str((pageIndex - 1) * pageSize) + "," + str(pageSize)
    sql = "select * from dm_user " + where + limit
    sqlcount = "select count(id) as countid from dm_user " + where
    get_dm_user = dm_user.get_by_sql(sql, sqlcount)
    if not get_dm_user:
        raise IdNotExist(f"系统中不存在 用户类ID 为 {productId} 的用户号结果.")
    return resp_200(data=get_dm_user, msg=f"查询到了 用户类ID 为 {productId} 的用户号结果.")

@router.post("/querybyproductid", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户号信息')
def query_dm_user_id(productId: int = 0, pageIndex: int = 1, pageSize: int = 10, db: Session = Depends(get_db)) -> Any:
    if productId != 0 :
        filter_query = (Dm_user.product_id == productId)
        get_dm_user = dm_user.get_by_page_filter(db, filter_query=filter_query, pageIndex=pageIndex, pageSize=pageSize )
        if not get_dm_user:
            raise IdNotExist(f"系统中不存在 用户类ID 为 {productId} 的用户号结果.")
        return resp_200(data=get_dm_user, msg=f"查询到了 用户类ID 为 {productId} 的用户号结果.")
    else:
        return resp_400()

@router.get("/querybloodsugarbyuserid", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户血糖信息')
def query_dm_user_bloodsugar_by_userid(
        *, dmuserid: str = "", pageIndex: int = 1, pageSize: int = 10,
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    if pageIndex<1:
        raise IdNotExist(f"页数小于1.")
    if pageSize<1:
        raise IdNotExist(f"每页记录数小于1.")
    order = " order by A.createtime desc"
    limit = " limit " + str((pageIndex - 1) * pageSize) + "," + str(pageSize)
    sql = "select A.id,A.dmuserid,A.dmusername,A.pmbg/10 as pmbg,A.pbg/10 as pbg,A.createtime,A.type,DU.deviceid from dm_pbloodsugar A left join dm_user DU on A.dmuserid=DU.id " + where + order + limit
    sqlcount = "select count(A.id) as countid from dm_pbloodsugar A left join dm_user DU on A.dmuserid=DU.id " + where
    get_bloodsugars = dm_user.get_by_sql(sql, sqlcount)
    if not get_bloodsugars:
        raise IdNotExist(f"系统中不存在 血糖查询关键字 用户ID为 {dmuserid} 的血糖结果.")
    count = get_bloodsugars['count']
    dataList = get_bloodsugars['dataList']
    #result = [dict(zip(result.keys(), result)) for result in dataList]
    for item in dataList:
        if int(item['type']) == 0:
            jielun1 = checkGluPMBGConclusion(item['pmbg'])
            jielun = jielun1['conclusion']
            hint = jielun1['hint']
        else:
            jielun1 = checkGluPBGConclusion(item['pbg'])
            jielun = jielun1['conclusion']
            hint = jielun1['hint']
        item['conclusion'] = jielun
        item['hint'] = hint
    back_data = {'count': count, 'dataList': dataList}
    return resp_200(data=back_data, msg=f"查询到了 血糖查询关键字 用户ID为 {dmuserid} 的血糖结果.")

@router.get("/querybloodsugarbyuserid2", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户血糖信息')
def query_dm_user_bloodsugar_by_userid2(
        dmuserid: str = "",
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    order = " order by A.id asc"
    sql = "select A.id,A.dmuserid,A.dmusername,A.pmbg/10 as pmbg,A.pbg/10 as pbg,A.createtime,DU.deviceid from dm_pbloodsugar A left join dm_user DU on A.dmuserid=DU.id " + where + order
    get_bloodsugars = dm_user.get_by_sql_no_count(sql)
    #print(sql)
    if not get_bloodsugars:
        raise IdNotExist(f"系统中不存在 血糖查询关键字 用户ID为 {dmuserid} 的血糖结果.")
    return resp_200(data=get_bloodsugars, msg=f"查询到了 血糖查询关键字 用户ID为 {dmuserid} 的血糖结果.")

@router.get("/querybloodpressurebyuserid", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户血压信息')
def query_dm_user_bloodpressure_by_userid(
        *, dmuserid: str = "", pageIndex: int = 1, pageSize: int = 10,
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    if pageIndex<1:
        raise IdNotExist(f"页数小于1.")
    if pageSize<1:
        raise IdNotExist(f"每页记录数小于1.")
    order = " order by A.createtime desc"
    limit = " limit " + str((pageIndex - 1) * pageSize) + "," + str(pageSize)
    sql = "select A.id,A.dmuserid,A.dmusername,A.dbp as dbp,A.sbp as sbp,A.createtime,DU.deviceid from dm_pbloodpressure A left join dm_user DU on A.dmuserid=DU.id " + where + order + limit
    sqlcount = "select count(A.id) as countid from dm_pbloodpressure A left join dm_user DU on A.dmuserid=DU.id " + where
    get_bloodpressures = dm_user.get_by_sql(sql, sqlcount)
    if not get_bloodpressures:
        raise IdNotExist(f"系统中不存在 血压查询关键字 用户ID为 {dmuserid} 的血压结果.")
    count = get_bloodpressures['count']
    dataList = get_bloodpressures['dataList']
    for item in dataList:
        jielun = checkBloodPressureConclusion(int(item['sbp']), int(item['dbp']))
        item['conclusion'] = jielun['conclusion']
        item['hint'] = jielun['hint']
    back_data = {'count': count, 'dataList': dataList}
    return resp_200(data=back_data, msg=f"查询到了 血压查询关键字 用户ID为 {dmuserid} 的血压结果.")

@router.get("/querybloodpressurebyuserid2", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户血压信息')
def query_dm_user_bloodpressure_by_userid2(
        dmuserid: str = "",
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    order = " order by A.createtime asc"
    sql = "select A.id,A.dmuserid,A.dmusername,A.dbp as dbp,A.sbp as sbp,A.createtime,DU.deviceid from dm_pbloodpressure A left join dm_user DU on A.dmuserid=DU.id " + where + order
    get_bloodpressures = dm_user.get_by_sql_no_count(sql)
    if not get_bloodpressures:
        raise IdNotExist(f"系统中不存在 血压查询关键字 用户ID为 {dmuserid} 的血压结果.")
    return resp_200(data=get_bloodpressures, msg=f"查询到了 血压查询关键字 用户ID为 {dmuserid} 的血压结果.")

@router.get("/queryheartratebyuserid", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户心率信息')
def query_dm_user_heartrate_by_userid(
        *, dmuserid: str = "", pageIndex: int = 1, pageSize: int = 10,
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    if pageIndex<1:
        raise IdNotExist(f"页数小于1.")
    if pageSize<1:
        raise IdNotExist(f"每页记录数小于1.")
    order = " order by A.createtime desc"
    limit = " limit " + str((pageIndex - 1) * pageSize) + "," + str(pageSize)
    sql = "select A.id,A.dmuserid,A.dmusername,A.heartrateurl,A.heartrate,A.createtime,DU.deviceid from dm_pheartrate A left join dm_user DU on A.dmuserid=DU.id " + where + order + limit
    sqlcount = "select count(A.id) as countid from dm_pheartrate A left join dm_user DU on A.dmuserid=DU.id " + where
    get_heartrates = dm_user.get_by_sql(sql, sqlcount)
    if not get_heartrates:
        raise IdNotExist(f"系统中不存在 心率查询关键字 用户ID为 {dmuserid} 的心率结果.")
    count = get_heartrates['count']
    dataList = get_heartrates['dataList']
    for item in dataList:
        jielun = checkHeartrateConclusion(int(item['heartrate']))
        item['conclusion'] = jielun['conclusion']
    back_data = {'count': count, 'dataList': dataList}
    return resp_200(data=back_data, msg=f"查询到了 心率查询关键字 用户ID为 {dmuserid} 的心率结果.")

@router.get("/queryheartratebyuserid2", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户心率信息')
def query_dm_user_heartrate_by_userid2(
        dmuserid: str = "",
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    order = " order by A.id asc"
    sql = "select A.id,A.dmuserid,A.dmusername,A.heartrateurl,A.heartrate,A.createtime,DU.deviceid from dm_pheartrate A left join dm_user DU on A.dmuserid=DU.id " + where + order
    get_heartrates = dm_user.get_by_sql_no_count(sql)
    #print(sql)
    if not get_heartrates:
        raise IdNotExist(f"系统中不存在 心率查询关键字 用户ID为 {dmuserid} 的心率结果.")
    return resp_200(data=get_heartrates, msg=f"查询到了 心率查询关键字 用户ID为 {dmuserid} 的心率结果.")

@router.get("/queryheartratebyuserid3", response_model=ResultModel[Dm_userOut], summary='根据 ID条件 查询用户心率信息')
def query_dm_user_heartrate_by_userid3(
        heartrateid: str = "0",
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if heartrateid == "0":
        raise IdNotExist(f"没有ID.")
    if heartrateid != "0":
        where = where + " and A.id = " + heartrateid
    order = " order by A.id asc"
    sql = "select A.id,A.dmuserid,A.dmusername,A.heartrateurl,A.heartrate,A.createtime,DU.deviceid from dm_pheartrate A left join dm_user DU on A.dmuserid=DU.id " + where + order
    get_heartrates = dm_user.get_by_sql_no_count(sql)
    #print(sql)
    if not get_heartrates:
        raise IdNotExist(f"系统中不存在 心率查询关键字 用户ID为 {heartrateid} 的心率结果.")
    return resp_200(data=get_heartrates[0], msg=f"查询到了 心率查询关键字 用户ID为 {heartrateid} 的心率结果.")

@router.get("/querybloodoxygenbyuserid", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户血氧信息')
def query_dm_user_bloodoxygen_by_userid(
        *, dmuserid: str = "", pageIndex: int = 1, pageSize: int = 10,
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    if pageIndex<1:
        raise IdNotExist(f"页数小于1.")
    if pageSize<1:
        raise IdNotExist(f"每页记录数小于1.")
    order = " order by A.createtime desc"
    limit = " limit " + str((pageIndex - 1) * pageSize) + "," + str(pageSize)
    sql = "select A.id,A.dmuserid,A.dmusername,A.bloodoxygenurl,A.spo,A.pulse,A.createtime,DU.deviceid from dm_pbloodoxygen A left join dm_user DU on A.dmuserid=DU.id " + where + order + limit
    sqlcount = "select count(A.id) as countid from dm_pbloodoxygen A left join dm_user DU on A.dmuserid=DU.id " + where
    get_bloodoxygens = dm_user.get_by_sql(sql, sqlcount)
    if not get_bloodoxygens:
        raise IdNotExist(f"系统中不存在 血氧查询关键字 用户ID为 {dmuserid} 的血氧结果.")
    count = get_bloodoxygens['count']
    dataList = get_bloodoxygens['dataList']
    for item in dataList:
        jielun = checkOxygenSpoConclusion(int(item['spo']))
        jielun1 = checkOxygenPulseConclusion(int(item['pulse']))
        item['conclusion'] = jielun['conclusion'] + '/' + jielun1['conclusion']
        item['hint'] = jielun['hint'] + '/' + jielun1['hint']
    back_data = {'count': count, 'dataList': dataList}
    return resp_200(data=back_data, msg=f"查询到了 血氧查询关键字 用户ID为 {dmuserid} 的血氧结果.")

@router.get("/querybloodoxygenbyuserid2", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户血氧信息')
def query_dm_user_bloodoxygen_by_userid2(
        dmuserid: str = "",
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    order = " order by A.id asc"
    sql = "select A.id,A.dmuserid,A.dmusername,A.bloodoxygenurl,A.spo,A.pulse,A.createtime,DU.deviceid from dm_pbloodoxygen A left join dm_user DU on A.dmuserid=DU.id " + where + order
    get_bloodoxygens = dm_user.get_by_sql_no_count(sql)
    #print(sql)
    if not get_bloodoxygens:
        raise IdNotExist(f"系统中不存在 血氧查询关键字 用户ID为 {dmuserid} 的血氧结果.")
    return resp_200(data=get_bloodoxygens, msg=f"查询到了 血氧查询关键字 用户ID为 {dmuserid} 的血氧结果.")

@router.get("/querybloodoxygenbyuserid3", response_model=ResultModel[Dm_userOut], summary='根据 ID条件 查询用户血氧信息')
def query_dm_user_bloodoxygen_by_userid3(
        bloodoxygenid: str = "0",
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if bloodoxygenid == "0":
        raise IdNotExist(f"没有ID.")
    if bloodoxygenid != "0":
        where = where + " and A.id = " + bloodoxygenid
    order = " order by A.id asc"
    sql = "select A.id,A.dmuserid,A.dmusername,A.bloodoxygenurl,A.spo,A.pulse,A.createtime,DU.deviceid from dm_pbloodoxygen A left join dm_user DU on A.dmuserid=DU.id " + where + order
    get_bloodoxygens = dm_user.get_by_sql_no_count(sql)
    #print(sql)
    if not get_bloodoxygens:
        raise IdNotExist(f"系统中不存在 血氧查询关键字 用户ID为 {bloodoxygenid} 的血氧结果.")
    return resp_200(data=get_bloodoxygens[0], msg=f"查询到了 血氧查询关键字 用户ID为 {bloodoxygenid} 的血氧结果.")

@router.get("/querytemperaturebyuserid", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户体温信息')
def query_dm_user_temperature_by_userid(
        *, dmuserid: str = "", pageIndex: int = 1, pageSize: int = 10,
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    if pageIndex<1:
        raise IdNotExist(f"页数小于1.")
    if pageSize<1:
        raise IdNotExist(f"每页记录数小于1.")
    order = " order by A.createtime desc"
    limit = " limit " + str((pageIndex - 1) * pageSize) + "," + str(pageSize)
    sql = "select A.id,A.dmuserid,A.dmusername,A.temperature,A.createtime,DU.deviceid from dm_ptemperature A left join dm_user DU on A.dmuserid=DU.id " + where + order + limit
    sqlcount = "select count(A.id) as countid from dm_ptemperature A left join dm_user DU on A.dmuserid=DU.id " + where
    get_temperatures = dm_user.get_by_sql(sql, sqlcount)
    if not get_temperatures:
        raise IdNotExist(f"系统中不存在 体温查询关键字 用户ID为 {dmuserid} 的体温结果.")
    count = get_temperatures['count']
    dataList = get_temperatures['dataList']
    for item in dataList:
        jielun = checkTemperatureConclusion(float(item['temperature']))
        item['conclusion'] = jielun['conclusion']
        item['hint'] = jielun['hint']
    back_data = {'count': count, 'dataList': dataList}
    return resp_200(data=back_data, msg=f"查询到了 体温查询关键字 用户ID为 {dmuserid} 的体温结果.")

@router.get("/querytemperaturebyuserid2", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户体温信息')
def query_dm_user_temperature_by_userid2(
        dmuserid: str = "",
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    order = " order by A.id asc"
    sql = "select A.id,A.dmuserid,A.dmusername,A.temperature,A.createtime,DU.deviceid from dm_ptemperature A left join dm_user DU on A.dmuserid=DU.id " + where + order
    get_temperatures = dm_user.get_by_sql_no_count(sql)
    if not get_temperatures:
        raise IdNotExist(f"系统中不存在 体温查询关键字 用户ID为 {dmuserid} 的体温结果.")
    return resp_200(data=get_temperatures, msg=f"查询到了 体温查询关键字 用户ID为 {dmuserid} 的体温结果.")

@router.get("/querymedicalrecordbyuserid", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户病历信息')
def query_dm_user_medicalrecord_by_userid(
        *, dmuserid: str = "", pageIndex: int = 1, pageSize: int = 10,
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    if pageIndex<1:
        raise IdNotExist(f"页数小于1.")
    if pageSize<1:
        raise IdNotExist(f"每页记录数小于1.")
    order = " order by A.createtime desc"
    limit = " limit " + str((pageIndex - 1) * pageSize) + "," + str(pageSize)
    sql = "select A.id,A.dmuserid,A.url,A.recordname,A.descr,A.createtime,A.deviceid,DU.`name` as dmusername from d_rhmedicalrecord A left join dm_user DU on A.dmuserid=DU.id " + where + order + limit
    sqlcount = "select count(A.id) as countid from d_rhmedicalrecord A left join dm_user DU on A.dmuserid=DU.id " + where
    get_medicalrecord = dm_user.get_by_sql(sql, sqlcount)
    if not get_medicalrecord:
        raise IdNotExist(f"系统中不存在 病历查询关键字 用户ID为 {dmuserid} 的病历结果.")
    return resp_200(data=get_medicalrecord, msg=f"查询到了 病历查询关键字 用户ID为 {dmuserid} 的病历结果.")

@router.get("/querymedicalrecordbyuserid2", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户病历信息')
def query_dm_user_medicalrecord_by_userid2(
        dmuserid: str = "",
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    order = " order by A.id asc"
    sql = "select A.id,A.dmuserid,DU.`name` as dmusername,A.url,A.recordname,A.createtime,A.descr,DU.deviceid from d_rhmedicalrecord A left join dm_user DU on A.dmuserid=DU.id " + where + order
    get_medicalrecord = dm_user.get_by_sql_no_count(sql)
    if not get_medicalrecord:
        raise IdNotExist(f"系统中不存在 病历查询关键字 用户ID为 {dmuserid} 的病历结果.")
    return resp_200(data=get_medicalrecord, msg=f"查询到了 病历查询关键字 用户ID为 {dmuserid} 的病历结果.")

@router.get("/querymedicalrecordbyuserid3", response_model=ResultModel[Dm_userOut], summary='根据 ID条件 查询用户病历信息')
def query_dm_user_medicalrecord_by_userid3(
        medicalrecordid: str = "0",
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if medicalrecordid == "0":
        raise IdNotExist(f"没有ID.")
    if medicalrecordid != "0":
        where = where + " and A.id = " + medicalrecordid
    order = " order by A.id asc"
    sql = "select A.id,A.dmuserid,DU.`name` as dmusername,A.url,A.recordname,A.descr,A.createtime,DU.deviceid from d_rhmedicalrecord A left join dm_user DU on A.dmuserid=DU.id " + where + order
    get_medicalrecords = dm_user.get_by_sql_no_count(sql)
    #print(sql)
    if not get_medicalrecords:
        raise IdNotExist(f"系统中不存在 病历查询关键字 用户ID为 {medicalrecordid} 的病历结果.")
    return resp_200(data=get_medicalrecords[0], msg=f"查询到了 病历查询关键字 用户ID为 {medicalrecordid} 的病历结果.")

@router.get("/queryphysiquebyuserid", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户体质辨识')
def query_dm_user_physique_by_userid(
        *, dmuserid: str = "", pageIndex: int = 1, pageSize: int = 10,
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if dmuserid != "":
        where = where + " and A.dmuserid = " + dmuserid
    if pageIndex<1:
        raise IdNotExist(f"页数小于1.")
    if pageSize<1:
        raise IdNotExist(f"每页记录数小于1.")
    order = " order by A.createtime desc"
    limit = " limit " + str((pageIndex - 1) * pageSize) + "," + str(pageSize)
    sql = "select A.id,A.dmuserid,A.dmusername,A.physique,A.createtime,DU.deviceid from dm_physique A left join dm_user DU on A.dmuserid=DU.id " + where + order + limit
    sqlcount = "select count(A.id) as countid from dm_physique A left join dm_user DU on A.dmuserid=DU.id " + where
    get_physique = dm_user.get_by_sql(sql, sqlcount)
    if not get_physique:
        raise IdNotExist(f"系统中不存在 体质辨识关键字 用户ID为 {dmuserid} 的体质辨识结果.")
    count = get_physique['count']
    dataList = get_physique['dataList']
    for item in dataList:
        jielun1 = checkPhysiqueConclusion(item['physique'])
        jielunphysique = jielun1['colorstring']
        hint = jielun1['hint']
        jielun = jielun1['conclusion']
        item['physique'] = jielunphysique
        item['hint'] = hint
        item['conclusion'] = jielun
    back_data = {'count': count, 'dataList': dataList}
    return resp_200(data=back_data, msg=f"查询到了 查询关键字 用户ID为 {dmuserid} 的体质辨识结果.")

@router.get("/queryphysiquebyuserid2", response_model=ResultModel[Dm_userOut], summary='根据 条件 查询用户体质辨识')
def query_dm_user_physique_by_userid2(
        physiqueid: str = "0",
        db: Session = Depends(get_db), current_user: Admin = Depends(get_current_user)
    ) -> Any:
    where = " where 1=1 "
    strWhereUser = getUserWhere(db,current_user)
    where += strWhereUser
    if physiqueid == "0":
        raise IdNotExist(f"没有ID.")
    if physiqueid != "0":
        where = where + " and A.id = " + physiqueid
    order = " order by A.id asc"
    sql = "select A.id,A.dmuserid,A.dmusername,A.physique,A.createtime,DU.deviceid from dm_physique A left join dm_user DU on A.dmuserid=DU.id " + where + order
    get_heartrates = dm_user.get_by_sql_fetchone1(sql)
    if len(get_heartrates) < 1:
        raise IdNotExist(f"系统中不存在 体质辨识查询关键字 用户ID为 {physiqueid} 的体质辨识结果.")
    dict_data = get_heartrates[0]
    jielunphysique = ''
    hint = ''
    jielun = ''
    for key,value in dict_data.items():
        if (key == 'physique'):
            jielun1 = checkPhysiqueConclusion(value)
            jielunphysique = jielun1['colorstring']
            hint = jielun1['hint']
            jielun = jielun1['conclusion']
    dict_data['physique'] = jielunphysique
    dict_data['hint'] = hint
    dict_data['conclusion'] = jielun
    return resp_200(data=dict_data, msg=f"查询到了 体质辨识查询关键字 用户ID为 {physiqueid} 的体质辨识结果.")

@router.get("/outexcel", response_model=ResultModel[Dm_userOut], summary='根据 条件 导出Excel信息')
def user_outexcel(idstr: str = '', pageIndex: int = 1, pageSize: int = 1000) -> Any:
    idary = idstr.split(',')
    where = " where 1=1"
    limit = " limit " + str((pageIndex - 1) * pageSize) + "," + str(pageSize)
    order = " order by A.id asc"
    if len(idary)>0:
        a = []
        for index, num in enumerate(idary):
            mydataid = mydataname = mydataphone = mydataage = mydatasex = mydataheight = mydataweight = mydatadeviceid = ''
            mydatadevicename = mydataidcard = mydatacityid = mydatacityname = mydatabloodtypename = mydataallergyname = mydatamedicalname = ''
            mydatageneticname = mydatacreatetime = mydatadescr = ''
            wb = Workbook()
            # 居中样式
            align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # 边框样式
            border = Border(left=Side(border_style='thin'),
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'),
                            bottom=Side(border_style='thin'))
            where1 = " and A.id = " + str(num)
            sql = "select A.*,DV.name as devicename from dm_user A left join device DV on A.deviceid = DV.device_code " + where + where1# + order + limit#查找用户表
            get_user_data = dm_user.get_by_sql_no_count(sql)
            if not get_user_data:
                ws = wb.active
                ws.title = str(num) + "用户信息"
                ws.merge_cells('A1:O1')
                ws['A1'] = str(num) + "用户信息不存在"
                t = time.time()
                nowtime = int(round(t * 1000))
                filename = f"{settings.STATIC_DIR}/uploads/{str(num)}_{str(nowtime)}.xlsx"
                wb.save(filename)
                a.append(filename)
                continue
                #raise IdNotExist(f"系统中不存在 ID 为 {num} 的用户.")
            else:
                for result in get_user_data:
                    for k, v in result.items():
                        if k == "id" :
                            mydataid = v
                        elif k == "name" :
                            if v != '' and v != None:
                                mydataname = v
                        elif k == "headicon" :
                            if v != '' and v != None:
                                mydataheadicon = v
                        elif k == "phone" :
                            if v != '' and v != None:
                                mydataphone = v
                        elif k == "age" :
                            if v != '' and v != None:
                                mydataage = v
                        elif k == "sex" :
                            if v != '' and v != None:
                                if int(v) == 1:
                                    mydatasex = '女'
                                else:
                                    mydatasex = '男'
                        elif k == "height" :
                            if v != '' and v != None:
                                mydataheight = v
                        elif k == "weight" :
                            if v != '' and v != None:
                                mydataweight = v
                        elif k == "deviceid" :
                            if v != '' and v != None:
                                mydatadeviceid = v
                        elif k == "devicename" :
                            if v != '' and v != None:
                                mydatadevicename = v
                        elif k == "idcard" :
                            if v != '' and v != None:
                                mydataidcard = v
                        elif k == "cityid" :
                            if v != '' and v != None:
                                mydatacityid = v
                        elif k == "cityname" :
                            if v != '' and v != None:
                                mydatacityname = v
                        elif k == "bloodtypename" :
                            if v != '' and v != None:
                                mydatabloodtypename = v
                        elif k == "allergyname" :
                            if v != '' and v != None:
                                mydataallergyname = v
                        elif k == "medicalname" :
                            if v != '' and v != None:
                                mydatamedicalname = v
                        elif k == "geneticname" :
                            if v != '' and v != None:
                                mydatageneticname = v
                        elif k == "createtime" :
                            if v != '' and v != None:
                                mydatacreatetime = v
                        elif k == "descr" :
                            if v != '' and v != None:
                                mydatadescr = v

                ws = wb.active
                row1 = [
                    ['用户ID', str(mydataid), '', '用户姓名', str(mydataname), '', '联系电话', str(mydataphone), '', '年龄', str(mydataage) + '岁', '', '性别', str(mydatasex)],
                    ['身高', str(mydataheight) + 'CM', '', '体重', str(mydataweight) + 'KG', '', '设备串号', str(mydatadeviceid), '', '设备名称', str(mydatadevicename), '', '身份证', str(mydataidcard)],
                    ['城市编号', str(mydatacityid), '', '城市名称', str(mydatacityname), '', '血型', str(mydatabloodtypename), '', '过敏史', str(mydataallergyname), '', '既往病史', str(mydatamedicalname)],
                    ['遗传史', str(mydatageneticname), '', '创建时间', str(mydatacreatetime), '', '健康说明', str(mydatadescr)]
                ]
                #print(row1)
                #ws1 = wb.create_sheet()
                #ws2 = wb.create_sheet(0)
                ws.title = str(mydataname) + "用户信息"
                #ws4 = wb.get_sheet_by_name(str(mydataname) + "用户信息")
                # 合并单元格A1-K1
                ws.merge_cells('A1:O1')
                ws['A1'] = str(mydataname) + "用户信息"
                # 设置font 字体
                font_title = Font(u'微软雅黑', size=18)
                # 引用font字体
                ws['A1'].font = font_title
                # 引用居中样式
                ws['A1'].alignment = align
                # 设置行高 第一行 40
                ws.row_dimensions[1].height = 40
                for row in row1:
                    ws.append(row)

                ws.merge_cells('B2:C2')
                ws.merge_cells('E2:F2')
                ws.merge_cells('H2:I2')
                ws.merge_cells('K2:L2')
                ws.merge_cells('N2:O2')
                ws.merge_cells('B3:C3')
                ws.merge_cells('E3:F3')
                ws.merge_cells('H3:I3')
                ws.merge_cells('K3:L3')
                ws.merge_cells('N3:O3')
                ws.merge_cells('B4:C4')
                ws.merge_cells('E4:F4')
                ws.merge_cells('H4:I4')
                ws.merge_cells('K4:L4')
                ws.merge_cells('N4:O4')
                ws.merge_cells('B5:C5')
                ws.merge_cells('E5:F5')
                ws.merge_cells('H5:I5')
                ws.merge_cells('K5:L5')
                ws.merge_cells('N5:O5')
                for row in ws.iter_rows(min_row=2, max_row=5, max_col=15):
                    i = 0
                    for cell in row:
                        cell.alignment = align
                        cell.border = border
                        i += 1

            #===================================血糖======================================
            ws1 = wb.create_sheet()
            ws1.title = "血糖信息"
            # 合并单元格A1-K1
            ws1.merge_cells('A1:I1')
            ws1['A1'] = str(mydataname) + "血糖信息"
            # 设置font 字体
            font_title = Font(u'微软雅黑', size=18)
            # 引用font字体
            ws1['A1'].font = font_title
            # 引用居中样式
            ws1['A1'].alignment = align
            # 设置行高 第一行 40
            ws1.row_dimensions[1].height = 40
            row1 = [
                ['用户姓名', str(mydataname), '', '联系电话', str(mydataphone), '', '年龄', str(mydataage) + '岁', ''],
                ['性别', str(mydatasex), '', '身高', str(mydataheight) + 'CM', '', '体重', str(mydataweight) + 'KG', '']
            ]
            firstrow = 1
            for row in row1:
                firstrow += 1
                ws1.append(row)
            for i in range(2, firstrow + 1):#合并单元格
                ws1.merge_cells('B' + str(i) + ':C' + str(i))
                ws1.merge_cells('E' + str(i) + ':F' + str(i))
                ws1.merge_cells('H' + str(i) + ':I' + str(i))
            for row in ws1.iter_rows(min_row=2, max_row=3, max_col=9):#设置边框设置居中
                i = 0
                for cell in row:
                    cell.alignment = align
                    cell.border = border
                    i += 1
            #--------------------------------餐前血糖---------------------------------------------
            ws1.append(['餐前血糖'])
            # 合并单元格A1-K1
            ws1.merge_cells('A4:I4')
            # 设置font 字体
            font_title = Font(u'微软雅黑', size=12)
            # 引用font字体
            ws1['A4'].font = font_title
            # 引用居中样式
            ws1['A4'].alignment = align
            # 设置行高 第一行 40
            ws1.row_dimensions[4].height = 20
            where2 = " and A.dmuserid = " + str(num) + " and pmbg > 0"
            sql = "select A.id,A.dmuserid,A.dmusername,A.pmbg/10 as pmbg,A.pbg/10 as pbg,A.createtime from dm_pbloodsugar A " + where + where2 + order + limit  # 查找血糖表
            get_user_bloodsugar_pmbg_data = dm_user.get_by_sql_no_count(sql)
            row2 = [['序号','用户ID','用户姓名','餐前血糖','结论','','测量时间']]
            noData = False
            if not get_user_bloodsugar_pmbg_data:
                noData = True
            else:
                j = 0
                for result in get_user_bloodsugar_pmbg_data:
                    j = j + 1
                    dmuserid = dmusername = pmbg = createtime = ''
                    for k, v in result.items():
                        if k == "dmuserid" :
                            if v != '' and v != None:
                                dmuserid = v
                            else:
                                dmuserid = ''
                        elif k == "dmusername" :
                            if v != '' and v != None:
                                dmusername = v
                            else:
                                dmusername = ''
                        elif k == "pmbg" :
                            if v != '' and v != None:
                                pmbg = v
                            else:
                                pmbg = ''
                        elif k == "createtime" :
                            if v != '' and v != None:
                                createtime = v
                            else:
                                createtime = ''
                    jielun = checkGluPMBGConclusion(pmbg)
                    rowlist = [j, dmuserid, dmusername, pmbg, jielun['conclusion'], '', createtime]
                    row2.append(rowlist)
            firstrow = 4
            for row in row2:
                firstrow += 1
                ws1.append(row)
            for i in range(5, firstrow + 1):#合并测量时间
                ws1.merge_cells('E' + str(i) + ':F' + str(i))
                ws1.merge_cells('G' + str(i) + ':I' + str(i))
            for row in ws1.iter_rows(min_row = 5, max_row = firstrow, max_col = 9):#设置边框设置居中
                i = 0
                for cell in row:
                    cell.alignment = align
                    cell.border = border
                    i += 1

            # 创建图表
            c1 = LineChart()
            c1.title = "餐前血糖"  # 标题
            c1.style = 13  # 样式
            c1.y_axis.title = '正常值范围：3.9~6.1mmol/L'  # Y轴
            c1.x_axis.title = '测量时间'  # X轴

            #print(firstrow)
            # 选择数据范围
            data = Reference(ws1, min_col=4, min_row=5, max_col=4, max_row=firstrow)
            categories = Reference(ws1, min_col=5, min_row=6, max_col=5, max_row=firstrow)
            # 折线图数据，y轴名称为第一行列名称
            c1.add_data(data, titles_from_data=True)
            c1.set_categories(categories)
            '''# 线条样式
            s0 = c1.series[0]
            s0.marker.symbol = "triangle"  # triangle为三角形标记， 可选circle、dash、diamond、dot、picture、plus、square、star、triangle、x、auto
            s0.marker.graphicalProperties.solidFill = "FF0000"  # 填充颜色
            s0.marker.graphicalProperties.line.solidFill = "0000FF"  # 边框颜色
            # s0.graphicalProperties.line.noFill = True  # 改为True则隐藏线条，但显示标记形状

            s1 = c1.series[1]
            s1.graphicalProperties.line.solidFill = "00AAAA"
            s1.graphicalProperties.line.dashStyle = "sysDot"  # 线条点状样式
            s1.graphicalProperties.line.width = 80000  # 线条大小，最大20116800EMUs

            s2 = c1.series[2]  # 采用默认设置
            s2.smooth = True  # 线条平滑'''
            #--------------------------------餐后血糖---------------------------------------------
            ws1.append(['餐后血糖'])
            # 合并单元格A1-K1
            ws1.merge_cells('A' + str(firstrow + 1) + ':I' + str(firstrow + 1))
            # 设置font 字体
            font_title = Font(u'微软雅黑', size=12)
            # 引用font字体
            ws1['A' + str(firstrow + 1)].font = font_title
            # 引用居中样式
            ws1['A' + str(firstrow + 1)].alignment = align
            # 设置行高 第一行 40
            ws1.row_dimensions[(firstrow + 1)].height = 20
            where2 = " and A.dmuserid = " + str(num) + " and pbg > 0"
            sql = "select A.id,A.dmuserid,A.dmusername,A.pmbg/10 as pmbg,A.pbg/10 as pbg,A.createtime from dm_pbloodsugar A " + where + where2 + order + limit  # 查找血糖表
            get_user_bloodsugar_pbg_data = dm_user.get_by_sql_no_count(sql)
            row2 = [['序号','用户ID','用户姓名','餐后血糖','结论','','测量时间']]
            noData = False
            if not get_user_bloodsugar_pbg_data:
                noData = True
            else:
                j = 0
                for result in get_user_bloodsugar_pbg_data:
                    j = j + 1
                    dmuserid = dmusername = pbg = createtime = ''
                    for k, v in result.items():
                        if k == "dmuserid" :
                            if v != '' and v != None:
                                dmuserid = v
                            else:
                                dmuserid = ''
                        elif k == "dmusername" :
                            if v != '' and v != None:
                                dmusername = v
                            else:
                                dmusername = ''
                        elif k == "pbg" :
                            if v != '' and v != None:
                                pbg = v
                            else:
                                pbg = ''
                        elif k == "createtime" :
                            if v != '' and v != None:
                                createtime = v
                            else:
                                createtime = ''
                    jielun = checkGluPMBGConclusion(pbg)
                    rowlist = [j, dmuserid, dmusername, pbg, jielun['conclusion'], '', createtime]
                    row2.append(rowlist)
            secondrow = firstrow + 1
            for row in row2:
                secondrow += 1
                ws1.append(row)
            for i in range(firstrow + 2, secondrow + 1):#合并测量时间
                ws1.merge_cells('E' + str(i) + ':F' + str(i))
                ws1.merge_cells('G' + str(i) + ':I' + str(i))
            for row in ws1.iter_rows(min_row = firstrow + 2, max_row = secondrow, max_col = 9):#设置边框设置居中
                i = 0
                for cell in row:
                    cell.alignment = align
                    cell.border = border
                    i += 1

            # 创建图表
            c2 = LineChart()
            c2.title = "餐后血糖"  # 标题
            c2.style = 13  # 样式
            c2.y_axis.title = '正常值范围：小于7.8mmol/L'  # Y轴
            c2.x_axis.title = '测量时间'  # X轴

            if secondrow == firstrow + 1:
                secondrow = secondrow + 1
            beginrow = firstrow + 2
            endrow = secondrow
            # 选择数据范围
            data2 = Reference(ws1, min_col=4, min_row=beginrow, max_col=4, max_row=endrow)#Y轴数据值
            if noData:#没有餐后血糖数据
                categories2 = Reference(ws1, min_col=5, min_row=beginrow + 1, max_col=5, max_row=endrow + 1)#Y轴标注数据
            else:
                categories2 = Reference(ws1, min_col=5, min_row=beginrow + 1, max_col=5, max_row=endrow)#X轴标注数据
            # 折线图数据，y轴名称为第一行列名称
            c2.add_data(data2, titles_from_data=True)
            c2.set_categories(categories2)

            ws1.add_chart(c1, "A" + str(secondrow + 3))  # 餐前血糖图表位置
            ws1.add_chart(c2, "A" + str(secondrow + 3 + 15))  # 餐后血糖图表位置
            #===================================血压======================================
            ws2 = wb.create_sheet()
            ws2.title = "血压信息"
            # 合并单元格A1-K1
            ws2.merge_cells('A1:I1')
            ws2['A1'] = str(mydataname) + "血压信息"
            # 设置font 字体
            font_title = Font(u'微软雅黑', size=18)
            # 引用font字体
            ws2['A1'].font = font_title
            # 引用居中样式
            ws2['A1'].alignment = align
            # 设置行高 第一行 40
            ws2.row_dimensions[1].height = 40
            row1 = [
                ['用户姓名', str(mydataname), '', '联系电话', str(mydataphone), '', '年龄', str(mydataage) + '岁', ''],
                ['性别', str(mydatasex), '', '身高', str(mydataheight) + 'CM', '', '体重', str(mydataweight) + 'KG', '']
            ]
            firstrow = 1
            for row in row1:
                firstrow += 1
                ws2.append(row)
            for i in range(2, firstrow + 1):#合并单元格
                ws2.merge_cells('B' + str(i) + ':C' + str(i))
                ws2.merge_cells('E' + str(i) + ':F' + str(i))
                ws2.merge_cells('H' + str(i) + ':I' + str(i))
            for row in ws2.iter_rows(min_row=2, max_row=3, max_col=9):#设置边框设置居中
                i = 0
                for cell in row:
                    cell.alignment = align
                    cell.border = border
                    i += 1
            #--------------------------------血压数据---------------------------------------------
            ws2.append(['血压数据'])
            # 合并单元格A1-K1
            ws2.merge_cells('A4:I4')
            # 设置font 字体
            font_title = Font(u'微软雅黑', size=12)
            # 引用font字体
            ws2['A4'].font = font_title
            # 引用居中样式
            ws2['A4'].alignment = align
            # 设置行高 第一行 40
            ws2.row_dimensions[4].height = 20
            where3 = " and A.dmuserid = " + str(num)
            sql = "select A.id,A.dmuserid,A.dmusername,A.dbp,A.sbp,A.createtime from dm_pbloodpressure A " + where + where3 + order + limit  # 查找血压表
            get_user_bloodpressure_data = dm_user.get_by_sql_no_count(sql)
            row2 = [['序号','用户ID','用户姓名','舒张压','收缩压','结论', '','测量时间']]
            noData = False
            if not get_user_bloodpressure_data:
                noData = True
            else:
                j = 0
                for result in get_user_bloodpressure_data:
                    j = j + 1
                    dmuserid = dmusername = createtime = ''
                    dbp = sbp = 0.00
                    for k, v in result.items():
                        if k == "dmuserid" :
                            if v != '' and v != None:
                                dmuserid = v
                            else:
                                dmuserid = ''
                        elif k == "dmusername" :
                            if v != '' and v != None:
                                dmusername = v
                            else:
                                dmusername = ''
                        elif k == "dbp" :
                            if v != '' and v != None:
                                dbp = int(v)
                            else:
                                dbp = ''
                        elif k == "sbp" :
                            if v != '' and v != None:
                                sbp = int(v)
                            else:
                                sbp = ''
                        elif k == "createtime" :
                            if v != '' and v != None:
                                createtime = v
                            else:
                                createtime = ''
                    jielun = checkBloodPressureConclusion(sbp, dbp)
                    rowlist = [j, dmuserid, dmusername, dbp, sbp, jielun['conclusion'], '', createtime]
                    row2.append(rowlist)
            firstrow = 4
            for row in row2:
                firstrow += 1
                ws2.append(row)
            for i in range(5, firstrow + 1):#合并测量时间
                ws2.merge_cells('F' + str(i) + ':G' + str(i))
                ws2.merge_cells('H' + str(i) + ':I' + str(i))
            for row in ws2.iter_rows(min_row = 5, max_row = firstrow, max_col = 9):#设置边框设置居中
                i = 0
                for cell in row:
                    cell.alignment = align
                    cell.border = border
                    i += 1
            ws2.column_dimensions['I'].width = 15

            # 创建图表
            c3 = LineChart()
            c3.title = "血压"  # 标题
            c3.style = 13  # 样式
            c3.y_axis.title = '正常值范围：收缩压：90~140mmHg，舒张压：60~90mmHg'  # Y轴
            c3.x_axis.title = '测量时间'  # X轴

            #print(firstrow)
            # 选择数据范围
            data1 = Reference(ws2, min_col=4, min_row=5, max_col=4, max_row=firstrow)
            data2 = Reference(ws2, min_col=5, min_row=5, max_col=5, max_row=firstrow)
            categories = Reference(ws2, min_col=6, min_row=6, max_col=6, max_row=firstrow)
            # 折线图数据，y轴名称为第一行列名称
            c3.add_data(data1, titles_from_data=True)
            c3.add_data(data2, titles_from_data=True)
            c3.set_categories(categories)

            ws2.add_chart(c3, "A" + str(firstrow + 3))  # 血压图表位置
            #===================================心率======================================
            ws3 = wb.create_sheet()
            ws3.title = "心率信息"
            # 合并单元格A1-K1
            ws3.merge_cells('A1:I1')
            ws3['A1'] = str(mydataname) + "心率信息"
            # 设置font 字体
            font_title = Font(u'微软雅黑', size=18)
            # 引用font字体
            ws3['A1'].font = font_title
            # 引用居中样式
            ws3['A1'].alignment = align
            # 设置行高 第一行 40
            ws3.row_dimensions[1].height = 40
            row1 = [
                ['用户姓名', str(mydataname), '', '联系电话', str(mydataphone), '', '年龄', str(mydataage) + '岁', ''],
                ['性别', str(mydatasex), '', '身高', str(mydataheight) + 'CM', '', '体重', str(mydataweight) + 'KG', '']
            ]
            firstrow = 1
            for row in row1:
                firstrow += 1
                ws3.append(row)
            for i in range(2, firstrow + 1):#合并单元格
                ws3.merge_cells('B' + str(i) + ':C' + str(i))
                ws3.merge_cells('E' + str(i) + ':F' + str(i))
                ws3.merge_cells('H' + str(i) + ':I' + str(i))
            for row in ws3.iter_rows(min_row=2, max_row=3, max_col=9):#设置边框设置居中
                i = 0
                for cell in row:
                    cell.alignment = align
                    cell.border = border
                    i += 1
            #--------------------------------心率数据---------------------------------------------
            ws3.append(['心率数据'])
            # 合并单元格A1-K1
            ws3.merge_cells('A4:I4')
            # 设置font 字体
            font_title = Font(u'微软雅黑', size=12)
            # 引用font字体
            ws3['A4'].font = font_title
            # 引用居中样式
            ws3['A4'].alignment = align
            # 设置行高 第一行 40
            ws3.row_dimensions[4].height = 20
            where3 = " and A.dmuserid = " + str(num)
            sql = "select A.id,A.dmuserid,A.dmusername,A.heartrate,A.heartrateurl,A.createtime from dm_pheartrate A " + where + where3 + order + limit  # 查找心率表
            get_user_heartrate_data = dm_user.get_by_sql_no_count(sql)
            row2 = [['序号','用户ID','用户姓名','心率值','结论', '','测量时间']]
            noData = False
            heartrateurlary = []
            heartrateurlstr = ''
            createtime = ''
            heartrate = ''
            if not get_user_heartrate_data:
                noData = True
            else:
                j = 0
                for result in get_user_heartrate_data:
                    j = j + 1
                    dmuserid = dmusername = createtime = heartrateurl = ''
                    heartrate = 0.00
                    for k, v in result.items():
                        if k == "dmuserid" :
                            if v != '' and v != None:
                                dmuserid = v
                            else:
                                dmuserid = ''
                        elif k == "dmusername" :
                            if v != '' and v != None:
                                dmusername = v
                            else:
                                dmusername = ''
                        elif k == "heartrate" :
                            if v != '' and v != None:
                                heartrate = int(v)
                            else:
                                heartrate = ''
                        elif k == "heartrateurl" :
                            if v != '' and v != None:
                                heartrateurl = v
                            else:
                                heartrateurl = ''
                        elif k == "createtime" :
                            if v != '' and v != None:
                                createtime = v
                            else:
                                createtime = ''
                    jielun = checkHeartrateConclusion(heartrate)
                    rowlist = [j, dmuserid, dmusername, heartrate, jielun['conclusion'], '', createtime]
                    heartrateurlstr = heartrateurl.replace('[','').replace(']','')
                    heartrateurlary = [int(x) for x in heartrateurlstr.split(",")]
                    row2.append(rowlist)
            firstrow = 4
            for row in row2:
                firstrow += 1
                ws3.append(row)
            for i in range(5, firstrow + 1):#合并测量时间
                ws3.merge_cells('E' + str(i) + ':F' + str(i))
                ws3.merge_cells('G' + str(i) + ':I' + str(i))
            for row in ws3.iter_rows(min_row = 5, max_row = firstrow, max_col = 9):#设置边框设置居中
                i = 0
                for cell in row:
                    cell.alignment = align
                    cell.border = border
                    i += 1

            # 创建图表
            c3 = LineChart()
            c3.title = "心率"  # 标题
            c3.style = 13  # 样式
            c3.y_axis.title = '正常值范围：50~120bpm'  # Y轴
            c3.x_axis.title = '测量时间'  # X轴

            #print(firstrow)
            # 选择数据范围
            data1 = Reference(ws3, min_col=4, min_row=5, max_col=4, max_row=firstrow)
            #data2 = Reference(ws3, min_col=5, min_row=5, max_col=5, max_row=firstrow)
            categories = Reference(ws3, min_col=5, min_row=6, max_col=5, max_row=firstrow)
            # 折线图数据，y轴名称为第一行列名称
            c3.add_data(data1, titles_from_data=True)
            #c3.add_data(data2, titles_from_data=True)
            c3.set_categories(categories)

            #--------------设置最后一个心率图数据--------------------
            ws3.cell(1, 10).value = 'X'
            ws3.cell(1, 11).value = '波形值'
            heartrateurendrow = 1
            if len(heartrateurlary) > 0:
                for i in range(0,len(heartrateurlary)):
                    heartrateurendrow += 1
                    ws3.cell(i + 2, 10).value = i + 1
                    ws3.cell(i + 2, 11).value = heartrateurlary[i]
            # 创建图表心率波形图
            c4 = LineChart()
            c4.title = str(createtime) + "心率波形图"  # 标题
            c4.style = 13  # 样式
            c4.y_axis.title = '心率值:' + str(heartrate) + '，心率波形'  # Y轴
            c4.x_axis.title = 'X轴'  # X轴
            c4.x_axis.tickLblPos = "low" #设置X轴标签在底部
            #c4.legend = None
            #print(firstrow)
            # 选择数据范围
            data1 = Reference(ws3, min_col=11, min_row=1, max_col=11, max_row=heartrateurendrow)
            #data2 = Reference(ws3, min_col=5, min_row=5, max_col=5, max_row=firstrow)
            categories = Reference(ws3, min_col=10, min_row=2, max_col=10, max_row=heartrateurendrow)
            # 折线图数据，y轴名称为第一行列名称
            c4.add_data(data1, titles_from_data=True)
            #c3.add_data(data2, titles_from_data=True)
            c4.set_categories(categories)

            ws3.add_chart(c3, "A" + str(firstrow + 3))  # 心率图表位置
            ws3.add_chart(c4, "A" + str(firstrow + 3 + 15))  # 心率图表位置
            #===================================血氧======================================
            ws4 = wb.create_sheet()
            ws4.title = "血氧信息"
            # 合并单元格A1-K1
            ws4.merge_cells('A1:I1')
            ws4['A1'] = str(mydataname) + "血氧信息"
            # 设置font 字体
            font_title = Font(u'微软雅黑', size=18)
            # 引用font字体
            ws4['A1'].font = font_title
            # 引用居中样式
            ws4['A1'].alignment = align
            # 设置行高 第一行 40
            ws4.row_dimensions[1].height = 40
            row1 = [
                ['用户姓名', str(mydataname), '', '联系电话', str(mydataphone), '', '年龄', str(mydataage) + '岁', ''],
                ['性别', str(mydatasex), '', '身高', str(mydataheight) + 'CM', '', '体重', str(mydataweight) + 'KG', '']
            ]
            firstrow = 1
            for row in row1:
                firstrow += 1
                ws4.append(row)
            for i in range(2, firstrow + 1):#合并单元格
                ws4.merge_cells('B' + str(i) + ':C' + str(i))
                ws4.merge_cells('E' + str(i) + ':F' + str(i))
                ws4.merge_cells('H' + str(i) + ':I' + str(i))
            for row in ws4.iter_rows(min_row=2, max_row=3, max_col=9):#设置边框设置居中
                i = 0
                for cell in row:
                    cell.alignment = align
                    cell.border = border
                    i += 1
            #--------------------------------血氧数据---------------------------------------------
            ws4.append(['血氧数据'])
            # 合并单元格A1-K1
            ws4.merge_cells('A4:I4')
            # 设置font 字体
            font_title = Font(u'微软雅黑', size=12)
            # 引用font字体
            ws4['A4'].font = font_title
            # 引用居中样式
            ws4['A4'].alignment = align
            # 设置行高 第一行 40
            ws4.row_dimensions[4].height = 20
            where4 = " and A.dmuserid = " + str(num)
            sql = "select A.id,A.dmuserid,A.dmusername,A.spo,A.pulse,A.bloodoxygenurl,A.createtime from dm_pbloodoxygen A " + where + where4 + order + limit  # 查找血氧表
            get_user_bloodoxygen_data = dm_user.get_by_sql_no_count(sql)
            row2 = [['序号','用户ID','用户姓名','饱和度','脉率','结论','','测量时间']]
            noData = False
            bloodoxygenurlary = []
            bloodoxygenurlstr = ''
            createtime = ''
            spo = ''
            pulse = ''
            if not get_user_bloodoxygen_data:
                noData = True
            else:
                j = 0
                for result in get_user_bloodoxygen_data:
                    j = j + 1
                    dmuserid = dmusername = createtime = bloodoxygenurl = ''
                    spo = pulse = 0.00
                    for k, v in result.items():
                        if k == "dmuserid" :
                            if v != '' and v != None:
                                dmuserid = v
                            else:
                                dmuserid = ''
                        elif k == "dmusername" :
                            if v != '' and v != None:
                                dmusername = v
                            else:
                                dmusername = ''
                        elif k == "spo" :
                            if v != '' and v != None:
                                spo = int(v)
                            else:
                                spo = ''
                        elif k == "pulse" :
                            if v != '' and v != None:
                                pulse = int(v)
                            else:
                                pulse = ''
                        elif k == "bloodoxygenurl" :
                            if v != '' and v != None:
                                bloodoxygenurl = v
                            else:
                                bloodoxygenurl = ''
                        elif k == "createtime" :
                            if v != '' and v != None:
                                createtime = v
                            else:
                                createtime = ''
                    jielun = checkOxygenSpoConclusion(spo)
                    jielun1 = checkOxygenPulseConclusion(pulse)
                    jielun2 = jielun['conclusion'] + "/" + jielun1['conclusion']
                    rowlist = [j, dmuserid, dmusername, spo, pulse, jielun2, '', createtime]
                    bloodoxygenurlstr = bloodoxygenurl.replace('[','').replace(']','')
                    bloodoxygenurlary = [int(x) for x in bloodoxygenurlstr.split(",")]
                    row2.append(rowlist)
            firstrow = 4
            for row in row2:#添加数据
                firstrow += 1
                ws4.append(row)
            for i in range(5, firstrow + 1):#合并测量时间
                ws4.merge_cells('F' + str(i) + ':G' + str(i))
                ws4.merge_cells('H' + str(i) + ':I' + str(i))
            for row in ws4.iter_rows(min_row = 5, max_row = firstrow, max_col = 9):#设置边框设置居中
                i = 0
                for cell in row:
                    cell.alignment = align
                    cell.border = border
                    i += 1
            ws4.column_dimensions['I'].width = 15

            # 创建图表
            c3 = LineChart()
            c3.title = "血氧"  # 标题
            c3.style = 13  # 样式
            c3.y_axis.title = '正常值范围：饱和度：94%以上，脉率：50~120bpm'  # Y轴
            c3.x_axis.title = '测量时间'  # X轴

            #print(firstrow)
            # 选择数据范围
            data1 = Reference(ws4, min_col=4, min_row=5, max_col=4, max_row=firstrow)
            data2 = Reference(ws4, min_col=5, min_row=5, max_col=5, max_row=firstrow)
            categories = Reference(ws4, min_col=6, min_row=6, max_col=6, max_row=firstrow)
            # 折线图数据，y轴名称为第一行列名称
            c3.add_data(data1, titles_from_data=True)
            c3.add_data(data2, titles_from_data=True)
            c3.set_categories(categories)

            #--------------设置最后一个血氧图数据--------------------
            ws4.cell(1, 10).value = 'X'
            ws4.cell(1, 11).value = '波形值'
            heartrateurendrow = 1
            if len(bloodoxygenurlary) > 0:
                for i in range(0,len(bloodoxygenurlary)):
                    heartrateurendrow += 1
                    ws4.cell(i + 2, 10).value = i + 1
                    ws4.cell(i + 2, 11).value = bloodoxygenurlary[i]
            # 创建图表血氧波形图
            c4 = LineChart()
            c4.title = str(createtime) + "血氧波形图"  # 标题
            c4.style = 13  # 样式
            c4.y_axis.title = '饱和度：' + str(spo) + '%，脉率：' + str(pulse) + '，血氧波形'  # Y轴
            c4.x_axis.title = 'X轴'  # X轴
            c4.x_axis.tickLblPos = "low" #设置X轴标签在底部
            #c4.legend = None
            #print(firstrow)
            # 选择数据范围
            data1 = Reference(ws4, min_col=11, min_row=1, max_col=11, max_row=heartrateurendrow)
            #data2 = Reference(ws4, min_col=5, min_row=5, max_col=5, max_row=firstrow)
            categories = Reference(ws4, min_col=10, min_row=2, max_col=10, max_row=heartrateurendrow)
            # 折线图数据，y轴名称为第一行列名称
            c4.add_data(data1, titles_from_data=True)
            #c3.add_data(data2, titles_from_data=True)
            c4.set_categories(categories)

            ws4.add_chart(c3, "A" + str(firstrow + 3))  # 血氧图表位置
            ws4.add_chart(c4, "A" + str(firstrow + 3 + 15))  # 血氧图表位置
            #===================================体温======================================
            ws5 = wb.create_sheet()
            ws5.title = "体温信息"
            # 合并单元格A1-K1
            ws5.merge_cells('A1:I1')
            ws5['A1'] = str(mydataname) + "体温信息"
            # 设置font 字体
            font_title = Font(u'微软雅黑', size=18)
            # 引用font字体
            ws5['A1'].font = font_title
            # 引用居中样式
            ws5['A1'].alignment = align
            # 设置行高 第一行 40
            ws5.row_dimensions[1].height = 40
            row1 = [
                ['用户姓名', str(mydataname), '', '联系电话', str(mydataphone), '', '年龄', str(mydataage) + '岁', ''],
                ['性别', str(mydatasex), '', '身高', str(mydataheight) + 'CM', '', '体重', str(mydataweight) + 'KG', '']
            ]
            firstrow = 1
            for row in row1:
                firstrow += 1
                ws5.append(row)
            for i in range(2, firstrow + 1):#合并单元格
                ws5.merge_cells('B' + str(i) + ':C' + str(i))
                ws5.merge_cells('E' + str(i) + ':F' + str(i))
                ws5.merge_cells('H' + str(i) + ':I' + str(i))
            for row in ws5.iter_rows(min_row=2, max_row=3, max_col=9):#设置边框设置居中
                i = 0
                for cell in row:
                    cell.alignment = align
                    cell.border = border
                    i += 1
            #--------------------------------体温数据---------------------------------------------
            ws5.append(['体温数据'])
            # 合并单元格A1-K1
            ws5.merge_cells('A4:I4')
            # 设置font 字体
            font_title = Font(u'微软雅黑', size=12)
            # 引用font字体
            ws5['A4'].font = font_title
            # 引用居中样式
            ws5['A4'].alignment = align
            # 设置行高 第一行 40
            ws5.row_dimensions[4].height = 20
            where3 = " and A.dmuserid = " + str(num)
            sql = "select A.id,A.dmuserid,A.dmusername,A.temperature,A.createtime from dm_ptemperature A " + where + where3 + order + limit  # 查找体温表
            get_user_temperature_data = dm_user.get_by_sql_no_count(sql)
            row2 = [['序号','用户ID','用户姓名','体温','结论','','测量时间']]
            noData = False
            if not get_user_temperature_data:
                noData = True
            else:
                j = 0
                for result in get_user_temperature_data:
                    j = j + 1
                    dmuserid = dmusername = createtime = ''
                    temperature = 0.00
                    for k, v in result.items():
                        if k == "dmuserid" :
                            if v != '' and v != None:
                                dmuserid = v
                            else:
                                dmuserid = ''
                        elif k == "dmusername" :
                            if v != '' and v != None:
                                dmusername = v
                            else:
                                dmusername = ''
                        elif k == "temperature" :
                            if v != '' and v != None:
                                temperature = float(v)
                            else:
                                temperature = 0.00
                        elif k == "createtime" :
                            if v != '' and v != None:
                                createtime = v
                            else:
                                createtime = ''
                    jielun = checkTemperatureConclusion(temperature)
                    rowlist = [j, dmuserid, dmusername, temperature, jielun['conclusion'], '', createtime]
                    row2.append(rowlist)
            firstrow = 4
            for row in row2:
                firstrow += 1
                ws5.append(row)
            for i in range(5, firstrow + 1):#合并测量时间
                ws5.merge_cells('E' + str(i) + ':F' + str(i))
                ws5.merge_cells('G' + str(i) + ':I' + str(i))
            for row in ws5.iter_rows(min_row = 5, max_row = firstrow, max_col = 9):#设置边框设置居中
                i = 0
                for cell in row:
                    cell.alignment = align
                    cell.border = border
                    i += 1

            # 创建图表
            c4 = LineChart()
            c4.title = "体温"  # 标题
            c4.style = 13  # 样式
            c4.y_axis.title = '正常值范围：36.0~37.5℃'  # Y轴
            c4.x_axis.title = '测量时间'  # X轴

            #print(firstrow)
            # 选择数据范围
            data1 = Reference(ws5, min_col=4, min_row=5, max_col=4, max_row=firstrow)
            #data2 = Reference(ws5, min_col=5, min_row=5, max_col=5, max_row=firstrow)
            categories = Reference(ws5, min_col=5, min_row=6, max_col=5, max_row=firstrow)
            # 折线图数据，y轴名称为第一行列名称
            c4.add_data(data1, titles_from_data=True)
            #c4.add_data(data2, titles_from_data=True)
            c4.set_categories(categories)

            ws5.add_chart(c4, "A" + str(firstrow + 3))  # 体温图表位置
            #===================================保存Excel======================================
            t = time.time()
            nowtime = int(round(t * 1000))
            filename = f"{settings.STATIC_DIR}/uploads/{str(mydataid)}_{str(nowtime)}.xlsx"
            wb.save(filename)
            a.append(filename)
        return resp_200(data=a, msg=f"查询到了遗传史结果.")
    else:
        raise IdNotExist(f"没有选择用户.")


